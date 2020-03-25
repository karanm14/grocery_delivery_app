import flask
from flask import Flask, request, render_template
import pandas as pd

import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import Input, Output, State
import dash_table9
import numpy as np
from flask import Flask
# app = Flask(__name__)

'''

for i in Order_Req['Items']:
    temp =  data[data['Item Name'] == i['Item Name']]
    if (i['Item Qty'] <= temp['Tab'].values[0]) & (i['Item Qty'] <= temp['Item Qty'].values[0]):
        print(i)
        data.loc[(data['Item Name'] == i['Item Name']), ['Item Qty']] = temp['Item Qty'] - i['Item Qty']
    print('updated for {}'.format(Order_Req['ID']))
print(data)

'''
external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']

server = Flask(__name__)
app = dash.Dash(__name__, server = server, external_stylesheets=external_stylesheets)

app.config['suppress_callback_exceptions'] = True
# food = pd.read_csv('inventory.csv',index_col='Item Code')
# server = app.server()
# data = food[['Item Name', 'Item Price']].to_dict('rows')
food = pd.read_excel('inventory.xlsx', index_col='Item Code', sheet_name='Food')
toiletry = pd.read_excel('inventory.xlsx', index_col='Item Code', sheet_name='Toiletry')
cart = pd.DataFrame(columns=['Item', 'Quantity', 'Price'])
order = []


dropzone = pd.read_excel('dropzone.xlsx',usecols=['Drop Point', 'Drop Time'])
#dropzone = pd.DataFrame()
def generate_table(order, df, max_rows=10):
    temp = {}
    for i in order:
        if i['category'] == 'food':
            temp['Item'] = i['item']
            temp['Quantity'] = i['quantity']
            temp['Price'] = food[food['Item Name'] == i['item']]['Item Price'].values[0]

        if i['category'] == 'toiletry':
            temp['Item'] = i['item']
            temp['Quantity'] = i['quantity']
            temp['Price'] = toiletry[toiletry['Item Name'] == i['item']]['Item Price'].values[0]

        temp['Total'] = temp["Price"] * temp['Quantity']
        df = df.append(temp, ignore_index=True)
    #print(df)
    return html.Table([
        html.Thead(
            html.Tr([html.Th(col) for col in df.columns])
        ),
        html.Tbody([
            html.Tr([
                html.Td(df.iloc[i][col]) for col in df.columns])
            for i in df.index
        ])
    ])


def append_df_to_excel(filename, df):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    """
    from openpyxl import load_workbook

    import pandas as pd

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # try to open an existing workbook
    writer.book = load_workbook(filename)
    sheet_name = 'Sheet1'
    startrow = writer.book[sheet_name].max_row

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=None)

    # save the workbook
    writer.save()
# food_order = []
# toiletry_order = []

page_1_layout = html.Div([

    html.H4("Items added in your cart"),
    html.Div(id='cart'),
    dcc.Tabs(id='tabs', children=[
        dcc.Tab(label='Food', children=[
            # html.Div([html.Div("Food Item - Price List for Canteen"),html.Div()])
            # dash_table.DataTable(
            #    id = 'price table',
            #    columns = [{'name' : i, 'id': i } for i in ['Item Name', 'Item Price', 'Tab']],
            #    data = food[['Item Name', 'Item Price', 'Tab']].to_dict('rows'),
            #    style_cell = {'maxWidth': '10px', 'width':'10px'},
            #    style_table = {'width': '500px'}
            # ),
            html.Div([
                dcc.ConfirmDialog(id='food-confirm'),
                html.Div(id='food-hidden-div', style={'display': 'none'}),
                html.Label('Select from list and enter the quantity'),
                dcc.Dropdown(
                    id='food-item-list',
                    options=[{'label': i, 'value': i} for i in food['Item Name']],
                    value='Wheat'
                ),
                # html.Label('Quantity'),
                html.Div(id='Food Price'),
                dcc.Dropdown(
                    id='food-quantity-list',
                    options=[{'label': i, 'value': i} for i in
                             np.arange(0, food[food['Item Name'] == 'Wheat']['Tab'].values[0])],
                    value = 1
                ),
                html.Button(id='food-submit-button', n_clicks=0, children='Submit')
            ])
        ]),

        dcc.Tab(label='Toiletry', children=[
            html.Div("Toiletry Item - Price List for Canteen"),
            # dash_table.DataTable(
            #    id = 'price table',
            #    columns = [{'name' : i, 'id': i } for i in ['Item Name', 'Item Price', 'Tab']],
            #    data = food[['Item Name', 'Item Price', 'Tab']].to_dict('rows'),
            #    style_cell = {'maxWidth': '10px', 'width':'10px'},
            #    style_table = {'width': '500px'}
            # ),
            html.Div([
                # html.Label('Select from list and enter the quantity'),
                dcc.ConfirmDialog(id='toiletry-confirm'),
                html.Div(id='toiletry-hidden-div', style={'display': 'none'}),
                dcc.Dropdown(
                    id='toiletry-item-list',
                    options=[{'label': i, 'value': i} for i in toiletry['Item Name']],
                    value='Shampoo'
                ),
                # html.Label('Quantity'),
                html.Div(id='Toiletry Price'),
                dcc.Dropdown(
                    id='toiletry-quantity-list',
                    options=[{'label': i, 'value': i} for i in
                             np.arange(0, toiletry[toiletry['Item Name'] == 'Shampoo']['Tab'].values[0])],
                    value=1
                ),
                html.Button(id='toiletry-submit-button', n_clicks=0, children='Submit')
            ])
        ])
    ]),
    html.Div([
        html.Div(children="Finished selecting? Place your final order"),
        html.Button(id='Place Order', n_clicks=0, children='Place Order',disabled=True),
        dcc.ConfirmDialog(id='order-confirm'),
    ])

], style={'width': '49%'})

page_2_layout = html.Div([
    html.H1('Confirm your details to place order'),
    html.Div(id='order-hidden-div',style={'display':'none'}),
    html.Div(id = 'page-2-content',children = "Please select your desired drop point and time"),
    dcc.Dropdown(id='zone-and-time',
        options = [{'label' : str((i,j)), 'value' : str((i,j))} for (i,j) in zip(dropzone['Drop Point'].to_list(),
                                                 dropzone['Drop Time'].to_list())]
    ),
    html.Label('Please enter your name: '),
    dcc.Input(id='name',value='',type='text'),
    html.Button(id='submit-button', n_clicks = 0, children ='Submit',disabled=True),
    dcc.ConfirmDialog(id='name and zone confirm')

], style={'width': '49%'})

app.layout = html.Div([
    #dcc.Location(id='url', refresh=False),
    html.Div(id='page-content', children = page_1_layout)
])


@app.callback(
    [Output('food-quantity-list', 'options'),
     Output('Food Price', 'children')],
    [Input('food-item-list', 'value')]
)
def set_food_quantity_dropdown(selected_item):
    temp = [np.arange(1, j + 1) for j in [food[food['Item Name'] == selected_item]['Tab'].values[0]]]
    options = [{'label': i, 'value': i} for i in temp[0]]
    price = food[food['Item Name'] == selected_item]['Item Price'].values[0]
    children = str(price) + ' price per unit'
    return (options, children)


@app.callback(
    [Output('food-confirm', 'message'),
     Output('food-confirm', 'displayed')],
    [Input('food-submit-button', 'n_clicks')],
    [State('food-item-list', 'value'),
     State('food-quantity-list', 'value')]
)
def confirm_food_order(n_clicks, food_item, qty):
    if n_clicks:
        message = "Are you sure you want to add {} units of {} to the cart?".format(qty, food_item)
        return message, True
    if n_clicks == 0:
        raise dash.exceptions.PreventUpdate


@app.callback(
    [Output('food-hidden-div', 'children')],
    [Input('food-confirm', 'submit_n_clicks')],
    [State('food-item-list', 'value'),
     State('food-submit-button', 'n_clicks'),
     State('food-quantity-list', 'value')]
)
def place_food_order(submit_n_clicks, food_item, n_clicks, qty):
    # print(submit_n_clicks,food_item, qty)
    # print("//////////")
    # print(food_order)
    if submit_n_clicks:
        order.append({'item': food_item, 'quantity': qty, 'category': 'food'})
        #print("//////")
        #print(order)
        #    print('3333333333')
        message = "{} units of {} item added to cart".format(str(qty), str(food_item))
        return [message]
    if (submit_n_clicks == 0) or (n_clicks == 0):
        #    print('4444444444')
        raise dash.exceptions.PreventUpdate


@app.callback(
    [Output('toiletry-quantity-list', 'options'),
     Output('Toiletry Price', 'children')],
    [Input('toiletry-item-list', 'value')]
)
def set_toiletry_quantity_dropdown(selected_item):
    temp = [np.arange(1, j + 1) for j in [toiletry[toiletry['Item Name'] == selected_item]['Tab'].values[0]]]
    options = [{'label': i, 'value': i} for i in temp[0]]
    price = toiletry[toiletry['Item Name'] == selected_item]['Item Price'].values[0]
    children = str(price) + ' price per unit'
    return (options, children)


@app.callback(
    [Output('toiletry-confirm', 'message'),
     Output('toiletry-confirm', 'displayed')],
    [Input('toiletry-submit-button', 'n_clicks')],
    [State('toiletry-item-list', 'value'),
     State('toiletry-quantity-list', 'value')]
)
def confirm_toiletry_order(n_clicks, toilet_item, qty):
    if n_clicks:
        message = "Are you sure you want to order {} units of {} to the cart?".format(qty, toilet_item)
        return message, True
    if n_clicks == 0:
        raise dash.exceptions.PreventUpdate


@app.callback(
    [Output('toiletry-hidden-div', 'children')],
    [Input('toiletry-confirm', 'submit_n_clicks')],
    [State('toiletry-item-list', 'value'),
     State('toiletry-submit-button', 'n_clicks'),
     State('toiletry-quantity-list', 'value')]
)
def place_toiletry_order(submit_n_clicks, toiletry_item, n_clicks, qty):
    # print(submit_n_clicks,food_item, qty)
    # print("//////////")
    # print(food_order)
    if submit_n_clicks:
        order.append({'item': toiletry_item, 'quantity': qty, 'category': 'toiletry'})
        #print(order)
        #    print('3333333333')
        message = "{} units of {} item added to cart".format(str(qty), str(toiletry_item))
        return [message]
    if (submit_n_clicks == 0) or (n_clicks == 0):
        #    print('4444444444')
        raise dash.exceptions.PreventUpdate


@app.callback(
    [Output('cart', 'children'),
     Output('Place Order', 'disabled')],
    # [Input('toiletry-confirm', 'submit_n_clicks'),
    #[Input('order-confirm', 'submit_n_clicks'),
    [Input('food-hidden-div', 'children'),
     Input('toiletry-hidden-div', 'children')]
)
def display_cart(confirm1, confirm2):
    #print("11111111111",snc)
    if confirm1 or confirm2:
        #print("value of confirm1 ", confirm1, "\n")
        #print("value of confirm2 ", confirm2, "\n")
        #print("value of order ", order, "\n")
        #print("value of cart ", cart, "\n")
        #print(generate_table(order, cart))
        return generate_table(order, cart), False
    else:
        raise dash.exceptions.PreventUpdate



'''
@app.callback(
    [Output('order-confirm','message'),
     Output('order-confirm','displayed')],
    [Input('food-hidden-div','children'),
     Input('toiletry-hidden-div','children'),
     Input('order-confirm','submit_n_clicks'),
     Input('Place Order','n_clicks')]
)
def confirm_order(msg1, msg2, snc, n_clicks):
    print("22222222222",snc)
    if (msg1 or msg2) and (n_clicks):
        message ="Do you want to place the order?"
        return message, True
    raise dash.exceptions.PreventUpdate
'''
@app.callback(
    [Output('order-confirm','message'),
     Output('order-confirm','displayed')],
     [Input('Place Order','n_clicks')]
)
def confirm_order(n_clicks):
    if n_clicks:
        message ="Do you want to place the order?"
        return message, True
    raise dash.exceptions.PreventUpdate

@app.callback(
    Output('page-content','children'),
    [Input('order-confirm','submit_n_clicks')]
)
def place_order(submit_n_clicks):
    if submit_n_clicks:
        return page_2_layout
    raise dash.exceptions.PreventUpdate

@app.callback(
    Output('submit-button','disabled'),
    [Input('name','value'),
     Input('zone-and-time','value')]
)
def enable_order(name, zone_time):
    if name and zone_time:
        return False
    else:
        raise dash.exceptions.PreventUpdate

@app.callback(
    [Output('name and zone confirm','message'),
     Output('name and zone confirm', 'displayed')],
    [Input('submit-button','n_clicks')],
    [State('name','value'),
     State('zone-and-time','value')]
)
def confirm_name_zone(n_clicks,name,z_t):
    if n_clicks:
        message = '{} are you sure you want to schedule your order for delivery at {} tomorrow'.format(name,z_t)
        return message, True
    raise dash.exceptions.PreventUpdate

@app.callback(
    Output('order-hidden-div','children'),
    [Input('name and zone confirm','submit_n_clicks')],
    [State('name','value'),
     State('zone-and-time','value')]
)
def update_order_inventory(submit_n_clicks,name,z_t):
    if submit_n_clicks:
        df = {}
        df['name'] = name

        print(name, z_t, order)
        message = 'Successful'
        return message

'''
def generate_table(order, df = cart, max_rows=10):
    temp = {}
    for i in order:
        print(i)
        if i['category'] == 'food':
            temp['Item'] = i['item']
            temp['Quantity'] = i['quantity']
            temp['Price'] = food[food['Item Name'] == i['item']]['Item Price'].values[0]

        if i['category'] == 'toiletry':
            temp['Item'] = i['item']
            temp['Quantity'] = i['quantity']
            temp['Price'] = toiletry[toiletry['Item Name'] == i['item']]['Item Price'].values[0]
        temp['Total'] = temp["Price"]*temp['Quantity']
        df = df.append(temp, ignore_index=True)

    return html.Table([
        html.Thead(
            html.Tr([html.Th(col) for col in df.columns])
        ),
        html.Tbody([
            html.Tr([
                html.Td(df.iloc[i][col]) for col in df.columns])
                for i in range(min(len(df), max_rows))
        ])
    ])


'''

if __name__ == '__main__':
    app.run_server(debug=True,threaded=True)